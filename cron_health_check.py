#!/usr/bin/env python3
"""Cron health check — detects missed jobs and reruns them.

Runs 3x daily (10:00, 15:00, 22:00) via cron. For each tracked job, checks
whether a successful run happened within its expected freshness window. If
not, reruns the job inline and updates state.

Tracked jobs:
- mundo daily_post  → catchup_state.last_post_date == today
- mundo engage      → catchup_state.last_engage_date == today AND engage.log mtime < 4h ago
- garmin update     → garmin_update.log mtime < 24h ago
- reddit post       → reddit_post.log mtime < 24h ago AND post stamp >= today 15:00
- reddit comment    → reddit_comment.log mtime < 6h ago after first 14:00 firing
"""
import os, sys, json, subprocess, datetime as dt
from pathlib import Path

HOME    = Path(os.path.expanduser("~"))
CFG     = HOME / ".config/mundo-bot"
LOGS    = HOME / "Library/Logs/mundo-bot"
PY      = "/usr/bin/python3"
STATE_F = CFG / "catchup_state.json"
HEALTH_LOG = LOGS / "health_check.log"

NOW = dt.datetime.now()
TODAY = NOW.date()
ICT_HOUR = NOW.hour


def log(msg):
    line = f"[{NOW.strftime('%Y-%m-%d %H:%M')}] {msg}"
    print(line)
    with open(HEALTH_LOG, "a") as f:
        f.write(line + "\n")


def file_age_h(path):
    p = Path(path)
    if not p.exists():
        return float("inf")
    return (NOW.timestamp() - p.stat().st_mtime) / 3600


def state():
    if STATE_F.exists():
        return json.loads(STATE_F.read_text())
    return {}


def write_state(d):
    STATE_F.write_text(json.dumps(d))


def run_chain(*scripts, log_file):
    """Run a chain of python scripts, redirecting stdout/stderr to log_file."""
    with open(log_file, "a") as f:
        for s in scripts:
            r = subprocess.run([PY, str(s)], stdout=f, stderr=subprocess.STDOUT, timeout=1800)
            if r.returncode != 0:
                return False
    return True


def check_garmin():
    age = file_age_h(LOGS / "garmin_update.log")
    if age > 24:
        log(f"⚠ garmin stale ({age:.1f}h) — rerunning")
        ok = run_chain(CFG / "refresh_token.py", CFG / "garmin_daily_update.py",
                       log_file=LOGS / "garmin_update.log")
        log(f"  garmin rerun {'✓' if ok else '✗'}")
        return ok
    log(f"✓ garmin fresh ({age:.1f}h)")
    return True


def check_mundo_post():
    s = state()
    last = s.get("last_post_date", "")
    if last == TODAY.isoformat():
        log(f"✓ mundo post done today ({last})")
        return True
    log(f"⚠ mundo post stale (last={last}) — rerunning")
    ok = run_chain(CFG / "refresh_token.py", CFG / "mundo_daily_post.py",
                   log_file=LOGS / "daily.log")
    if ok:
        s = state(); s["last_post_date"] = TODAY.isoformat(); write_state(s)
    log(f"  mundo post rerun {'✓' if ok else '✗'}")
    return ok


def check_mundo_engage():
    age = file_age_h(LOGS / "engage.log")
    s = state()
    last = s.get("last_engage_date", "")
    if last == TODAY.isoformat() and age < 4:
        log(f"✓ mundo engage fresh (last={last}, {age:.1f}h)")
        return True
    log(f"⚠ mundo engage stale (last={last}, age={age:.1f}h) — rerunning")
    ok = run_chain(CFG / "refresh_token.py", CFG / "mundo_engage.py",
                   log_file=LOGS / "engage.log")
    if ok:
        s = state(); s["last_engage_date"] = TODAY.isoformat(); write_state(s)
    log(f"  mundo engage rerun {'✓' if ok else '✗'}")
    return ok


def check_reddit_post():
    # Cron fires 15:00. Only check after 15:30 ICT.
    if ICT_HOUR < 15 or (ICT_HOUR == 15 and NOW.minute < 30):
        log("· reddit post window not yet (cron 15:00)")
        return True
    log_path = LOGS / "reddit_post.log"
    if not log_path.exists():
        log("⚠ reddit_post.log missing — checking token first")
    else:
        mtime = dt.datetime.fromtimestamp(log_path.stat().st_mtime)
        if mtime.date() == TODAY and mtime.hour >= 15:
            log(f"✓ reddit post fresh (mtime {mtime:%H:%M})")
            return True
        log(f"⚠ reddit post stale (last mtime {mtime:%Y-%m-%d %H:%M}) — checking token")
    # Check token before rerunning — skip if dead (saves wasted API call + log spam)
    token_check = subprocess.run([PY, str(CFG / "reddit_token_check.py")],
                                 capture_output=True, timeout=10)
    if token_check.returncode != 0:
        log("· reddit token dead — skip retry (user must re-login at reddit.com)")
        return True  # not a failure for cron health purposes
    with open(log_path, "a") as f:
        r = subprocess.run([PY, str(CFG / "reddit_post.py"), "--mode", "post"],
                           stdout=f, stderr=subprocess.STDOUT, timeout=1800)
    log(f"  reddit post rerun {'✓' if r.returncode == 0 else '✗'}")
    return r.returncode == 0


def check_reddit_comment():
    # Cron fires 14, 18, 22. Check that some firing today succeeded if past 14:30.
    if ICT_HOUR < 14 or (ICT_HOUR == 14 and NOW.minute < 30):
        log("· reddit comment window not yet (cron 14:00)")
        return True
    log_path = LOGS / "reddit_comment.log"
    if not log_path.exists():
        log("⚠ reddit_comment.log missing — rerunning")
    else:
        mtime = dt.datetime.fromtimestamp(log_path.stat().st_mtime)
        if mtime.date() == TODAY:
            log(f"✓ reddit comment touched today ({mtime:%H:%M})")
            return True
        log(f"⚠ reddit comment stale (last mtime {mtime:%Y-%m-%d %H:%M}) — rerunning")
    with open(log_path, "a") as f:
        r = subprocess.run([PY, str(CFG / "reddit_post.py"), "--mode", "comment"],
                           stdout=f, stderr=subprocess.STDOUT, timeout=1800)
    log(f"  reddit comment rerun {'✓' if r.returncode == 0 else '✗'}")
    return r.returncode == 0


def check_sprint_tracker_orphan_rows():
    """Verify every <td rowspan="N"> Topic cell has exactly N <tr> siblings.
    Orphan rows = data rows outside any rowspan group → table mis-renders.
    Documented rule: feedback_sprint_tracker_weekly_rotation.md §Orphan-row rule.
    Runs daily — cheap REST GET; flags structural rot.
    """
    try:
        import urllib.request, base64, re
        auth = base64.b64encode(b'thainlq:mo-api-9HEsX7pdU8wGCrds1dkucPVq').decode()
        req = urllib.request.Request(
            'https://confluence.zalopay.vn/rest/api/content/318790357?expand=body.storage',
            headers={'Authorization': f'Basic {auth}'}
        )
        body = json.loads(urllib.request.urlopen(req, timeout=15).read())['body']['storage']['value']
        # Find snapshot table
        ts = body.find('<table class="relative-table wrapped"', 5500)
        if ts < 0:
            log("· sprint tracker orphan: snapshot table not found — skip")
            return True
        te = body.find('</tbody></table>', ts) + len('</tbody></table>')
        table = body[ts:te]
        # Locate Topic rowspan cells + count <tr> in each group
        topic_pat = re.compile(r'<td rowspan="(\d+)"[^>]*><strong>([^<]+)</strong></td>', re.DOTALL)
        anchors = list(topic_pat.finditer(table))
        if not anchors:
            log("· sprint tracker orphan: no Topic rowspan cells — skip")
            return True
        problems = []
        # Each Topic group spans [<tr> containing this anchor] + next (N-1) <tr> siblings
        for i, m in enumerate(anchors):
            rs = int(m.group(1)); name = m.group(2)
            # Anchor sits inside a <tr>; find its <tr> start
            tr_start = table.rfind('<tr', 0, m.start())
            # Walk forward rs+1 occurrences of <tr (the anchor's tr counts as #1)
            pos = tr_start
            tr_positions = [tr_start]
            for _ in range(rs):
                nxt = table.find('<tr', pos + 3)
                if nxt < 0:
                    break
                tr_positions.append(nxt)
                pos = nxt
            if len(tr_positions) < rs:
                problems.append(f'{name}: rowspan={rs} but only {len(tr_positions)} <tr> found')
            # Check next <tr> after group: if next anchor is BEFORE it, gap exists (orphans between groups)
            if i + 1 < len(anchors):
                next_anchor_tr = table.rfind('<tr', 0, anchors[i+1].start())
                # Last <tr> of current group = tr_positions[-1], its </tr>
                last_tr_end = table.find('</tr>', tr_positions[-1]) + len('</tr>')
                # Count <tr> opens between last_tr_end and next_anchor_tr
                gap = table[last_tr_end:next_anchor_tr]
                orphan_trs = re.findall(r'<tr[^>]*>', gap)
                # Separator rows (colspan="N" only) don't count as orphan data rows
                # Simple heuristic: data row has multiple <td>; separator has colspan
                data_orphans = 0
                for sep_m in re.finditer(r'<tr[^>]*>(.*?)</tr>', gap, re.DOTALL):
                    if 'colspan=' not in sep_m.group(1):
                        data_orphans += 1
                if data_orphans:
                    problems.append(f'{name}→next: {data_orphans} orphan row(s) between groups')
        if problems:
            log(f'⚠ sprint tracker orphan rows detected: {"; ".join(problems)}')
            log('  → see feedback_sprint_tracker_weekly_rotation.md §Orphan-row rule')
            return False
        log('✓ sprint tracker: no orphan rows')
        return True
    except Exception as e:
        log(f'· sprint tracker orphan check failed: {e}')
        return True


def check_sprint_tracker_monday_rotation():
    """If today is Monday and Sprint Tracker page 318790357 still has previous week
    marked as CURRENT, log a reminder. Does NOT auto-rotate (manual review needed).
    Rule documented at: 07 - Claude Memory/feedback/feedback_sprint_tracker_weekly_rotation.md
    """
    # Only check on Mondays (weekday 0)
    if NOW.weekday() != 0:
        return True
    # Only after 09:00 ICT (give morning routines time)
    if ICT_HOUR < 9:
        return True
    try:
        import urllib.request, base64
        auth = base64.b64encode(b'thainlq:mo-api-9HEsX7pdU8wGCrds1dkucPVq').decode()
        req = urllib.request.Request(
            'https://confluence.zalopay.vn/rest/api/content/318790357?expand=body.storage',
            headers={'Authorization': f'Basic {auth}'}
        )
        body = json.loads(urllib.request.urlopen(req, timeout=15).read())['body']['storage']['value']
        # Extract CURRENT W column anchor date — pattern: "DD-MMM (Wn) — CURRENT"
        import re
        m = re.search(r'<strong>(?:Now &mdash; )?(\d{1,2})-([A-Z][a-z]{2}) \(W(\d+)\) &mdash; CURRENT</strong>', body)
        if not m:
            log("⚠ sprint tracker: no CURRENT W column found — manual review")
            return False
        cur_day, cur_mon, cur_wn = int(m.group(1)), m.group(2), int(m.group(3))
        # If today's date != CURRENT column's anchor date, rotation due
        month_map = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}
        cur_date = dt.date(TODAY.year, month_map[cur_mon], cur_day)
        if cur_date == TODAY:
            log(f"✓ sprint tracker: CURRENT = W{cur_wn} ({cur_date}) matches today")
            return True
        log(f"⚠ sprint tracker rotation due — CURRENT is W{cur_wn} ({cur_date}), today is Monday {TODAY}")
        log("  → see vault feedback_sprint_tracker_weekly_rotation.md for procedure")
        log("  → manual rotation needed (not auto — content review required)")
        return False
    except Exception as e:
        log(f"⚠ sprint tracker check failed: {e}")
        return True  # don't block other checks


def check_excel_report_structure():
    """Check ekyc_nfc_report_v20.xlsx trend table has VNeID blended into FLOW NFC SR + FLOW % SHARE groups.
    Rule (2026-05-12): new flows blend into existing groups, no standalone trio column.
    """
    try:
        import openpyxl
        path = os.path.expanduser('~/Desktop/ekyc_nfc_report_v20.xlsx')
        if not os.path.exists(path):
            log(f"⚠ excel report missing: {path}")
            return True
        wb = openpyxl.load_workbook(path, read_only=False)
        ws = wb['Report']
        # Check H24 = 'VNeID' (blended into FLOW NFC SR group)
        h24 = ws['H24'].value
        # Check L24 = 'VNeID%' (blended into FLOW % SHARE group)
        l24 = ws['L24'].value
        # Check merge E23:H23 exists (FLOW NFC SR group expanded to 4 cols)
        merges = {str(m) for m in ws.merged_cells.ranges}
        e_merge_ok = 'E23:H23' in merges
        i_merge_ok = 'I23:L23' in merges
        # Detect overlapping merges (Excel corruption source — 13-May lesson)
        mlist = list(ws.merged_cells.ranges)
        overlaps = []
        for i, m1 in enumerate(mlist):
            for m2 in mlist[i+1:]:
                if (m1.min_row <= m2.max_row and m1.max_row >= m2.min_row
                    and m1.min_col <= m2.max_col and m1.max_col >= m2.min_col):
                    overlaps.append((str(m1), str(m2)))
        if overlaps:
            log(f"⚠ excel report: {len(overlaps)} overlapping merges (corruption risk!)")
            for a, b in overlaps[:3]:
                log(f"    {a} <-> {b}")
            return False
        if h24 == 'VNeID' and l24 == 'VNeID%' and e_merge_ok and i_merge_ok:
            log("✓ excel report: VNeID blended + no merge overlaps")
            return True
        log(f"⚠ excel report structure drift — H24={h24!r} L24={l24!r} E_merge={e_merge_ok} I_merge={i_merge_ok}")
        log("  → see project_ekyc_report_design.md 'New Flow Blending Rule (2026-05-12)'")
        return False
    except Exception as e:
        log(f"⚠ excel report check failed: {e}")
        return True


def main():
    log(f"=== cron health check (hour={ICT_HOUR}) ===")
    results = {
        "garmin":         check_garmin(),
        "mundo_post":     check_mundo_post(),
        "mundo_engage":   check_mundo_engage(),
        "reddit_post":    check_reddit_post(),
        "reddit_comment": check_reddit_comment(),
        "sprint_tracker_monday": check_sprint_tracker_monday_rotation(),
        "sprint_tracker_orphan": check_sprint_tracker_orphan_rows(),
        "excel_report_structure": check_excel_report_structure(),
    }
    failed = [k for k, v in results.items() if not v]
    log(f"=== done · ok={len(results)-len(failed)}/{len(results)} · failed={failed or 'none'} ===\n")
    sys.exit(0 if not failed else 1)


if __name__ == "__main__":
    main()
